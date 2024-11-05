import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sys
import logging
from data_loader import load_and_prepare_data
import os
from datetime import datetime

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

def validate_columns(df):
    """Validate and ensure required columns exist"""
    required_columns = {
        'posteam': ['possession_team', 'pos_team', 'off_team', 'team'],
        'defteam': ['defensive_team', 'def_team', 'defense'],
        'yardline_100': ['yards_to_goal', 'yard_line', 'yards_from_goal'],
        'play_type': ['play', 'type'],
        'play_id': ['play_number', 'play_sequence'],
        'yards_gained': ['yards', 'gain'],
        'down': ['current_down'],
        'ydstogo': ['yards_to_go', 'distance'],
        'game_id': ['game_code', 'gameid'],
        'drive': ['drive_id', 'drive_number']
    }

    # Remove field_goal_attempt and game_seconds_remaining from required validation
    # since they're created during data preparation

    missing_columns = []
    for col, alternatives in required_columns.items():
        if col not in df.columns:
            found = False
            for alt in alternatives:
                if alt in df.columns:
                    print(f"Using '{alt}' for required column '{col}'")
                    df[col] = df[alt]
                    found = True
                    break
            if not found:
                missing_columns.append(col)

    if missing_columns:
        print("\nMissing required columns:")
        for col in missing_columns:
            print(f"- {col} (alternatives: {', '.join(required_columns[col])})")
        print("\nAvailable columns in dataset:")
        print(list(df.columns))
        raise ValueError("Missing required columns")

    return df

def export_to_excel(results, team1, team2, output_dir="output"):
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Create filename with teams and date
    filename = f"{team1}_vs_{team2}_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Matchup Analysis"
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    subheader_fill = PatternFill(start_color="D0D3D4", end_color="D0D3D4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write title
    ws['A1'] = f"{team1} vs {team2} Matchup Analysis"
    ws.merge_cells('A1:E1')
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Define metric categories
    categories = [
        ("Drive Success Metrics", [
            ('Total Drives', 'total_drives'),
            ('Scoring Position Rate', 'scoring_position_rate', '%'),
            ('Touchdown Rate', 'touchdown_rate', '%'),
            ('Field Goal Attempt Rate', 'field_goal_attempt_rate', '%')
        ]),
        ("Red Zone Efficiency", [
            ('Red Zone Attempts', 'red_zone_attempts'),
            ('Red Zone Success Rate', 'red_zone_success_rate', '%')
        ]),
        ("Third Down Efficiency", [
            ('Third Down Attempts', 'third_down_attempts'),
            ('Third Down Conversion Rate', 'third_down_conversion_rate', '%')
        ]),
        ("Drive Metrics", [
            ('Avg Plays per Drive', 'avg_plays_per_drive'),
            ('Avg Yards per Drive', 'avg_yards_per_drive'),
            ('Avg Starting Position', 'avg_start_position'),
            ('Avg Time per Drive (sec)', 'avg_time_per_drive')
        ]),
        ("Turnover Metrics", [
            ('Turnover Rate', 'turnover_rate', '%')
        ])
    ]
    
    # Write column headers
    headers = ['Metric', f'{team1} Offense', f'{team1} Defense', f'{team2} Offense', f'{team2} Defense']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    current_row = 4
    
    # Write data by category
    for category, metrics in categories:
        # Write category header
        ws.cell(row=current_row, column=1, value=category).fill = subheader_fill
        ws.merge_cells(f'A{current_row}:E{current_row}')
        current_row += 1
        
        # Write metrics
        for metric_name, metric_key, *format_args in metrics:
            row = [
                metric_name,
                results[team1]['offense'][metric_key],
                results[team1]['defense'][metric_key],
                results[team2]['offense'][metric_key],
                results[team2]['defense'][metric_key]
            ]
            
            for col, value in enumerate(row, 1):
                cell = ws.cell(row=current_row, column=col)
                if col > 1:  # Format numbers
                    if format_args and '%' in format_args:
                        cell.value = f"{value:.1f}%"
                    else:
                        cell.value = f"{value:.1f}" if isinstance(value, float) else value
                else:
                    cell.value = value
                cell.border = border
            
            current_row += 1
        
        current_row += 1  # Add space between categories
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(filepath)
    print(f"\nAnalysis exported to {filepath}")

def analyze_matchup(df, team1, team2):
    """
    Analyze offensive and defensive performance metrics for two teams.
    
    Args:
        df (DataFrame): The main dataframe containing play-by-play data
        team1 (str): First team's abbreviation
        team2 (str): Second team's abbreviation
    
    Returns:
        dict: Dictionary containing analysis results for both teams
    """
    
    # Validate team names
    all_teams = set(df['posteam'].unique()) | set(df['defteam'].unique())
    if team1 not in all_teams:
        sys.exit(f"Error: Team '{team1}' not found in dataset")
    if team2 not in all_teams:
        sys.exit(f"Error: Team '{team2}' not found in dataset")
    
    # Validate columns first
    df = validate_columns(df)
    
    def calculate_team_metrics(team, as_offense=True):
        # Filter for the team's offensive or defensive drives
        team_role = 'posteam' if as_offense else 'defteam'
        team_drives = df[df[team_role] == team].copy()
        
        # Group by game_id and drive for drive-level metrics
        drive_outcomes = team_drives.groupby(['game_id', 'drive']).agg({
            'yardline_100': ['min', 'first'],  # Closest they got & starting position
            'play_type': lambda x: 'touchdown' in x.values,  # Did drive end in TD
            'field_goal_attempt': 'sum',  # Number of FG attempts
            'play_id': 'count',  # Plays per drive
            'yards_gained': 'sum',  # Yards per drive
            'game_seconds_remaining': lambda x: x.iloc[0] - x.iloc[-1],  # Time of possession
            'turnover': 'sum'  # Turnovers
        }).reset_index()
        
        # Calculate red zone opportunities and success
        red_zone_drives = drive_outcomes[drive_outcomes['yardline_100']['min'] <= 20]
        red_zone_success = red_zone_drives[drive_outcomes['play_type']]
        
        # Calculate third down metrics
        third_downs = team_drives[team_drives['down'] == 3]
        third_down_conversions = third_downs[
            (third_downs['yards_gained'] >= third_downs['ydstogo']) |
            (third_downs['first_down'] == 1)
        ]
        
        # Calculate drive-level metrics
        total_drives = len(drive_outcomes)
        scoring_position_drives = len(drive_outcomes[drive_outcomes['yardline_100']['min'] <= 20])
        touchdown_drives = len(drive_outcomes[drive_outcomes['play_type']])
        fg_attempt_drives = len(drive_outcomes[drive_outcomes['field_goal_attempt'] > 0])
        
        # Calculate averages
        avg_plays_per_drive = drive_outcomes['play_id'].mean()
        avg_yards_per_drive = drive_outcomes['yards_gained'].mean()
        avg_start_position = drive_outcomes['yardline_100']['first'].mean()
        avg_time_per_drive = drive_outcomes['game_seconds_remaining'].mean()
        
        return {
            'total_drives': total_drives,
            'scoring_position_rate': scoring_position_drives / total_drives * 100,
            'touchdown_rate': touchdown_drives / total_drives * 100,
            'field_goal_attempt_rate': fg_attempt_drives / total_drives * 100,
            
            # Red Zone Efficiency
            'red_zone_attempts': len(red_zone_drives),
            'red_zone_success_rate': len(red_zone_success) / len(red_zone_drives) * 100 if len(red_zone_drives) > 0 else 0,
            
            # Third Down Efficiency
            'third_down_attempts': len(third_downs),
            'third_down_conversion_rate': len(third_down_conversions) / len(third_downs) * 100 if len(third_downs) > 0 else 0,
            
            # Drive Metrics
            'avg_plays_per_drive': avg_plays_per_drive,
            'avg_yards_per_drive': avg_yards_per_drive,
            'avg_start_position': avg_start_position,
            'avg_time_per_drive': avg_time_per_drive,
            
            # Turnover Metrics
            'turnover_rate': drive_outcomes['turnover'].sum() / total_drives * 100
        }
    
    # Calculate metrics for both teams
    results = {
        team1: {
            'offense': calculate_team_metrics(team1, as_offense=True),
            'defense': calculate_team_metrics(team1, as_offense=False)
        },
        team2: {
            'offense': calculate_team_metrics(team2, as_offense=True),
            'defense': calculate_team_metrics(team2, as_offense=False)
        }
    }
    
    return results

def main():
    # Load data using the data_loader
    logging.info("Loading NFL play-by-play data...")
    df = load_and_prepare_data()
    
    if df is None:
        sys.exit("Failed to load data")
    
    # Parse command line arguments
    import argparse
    parser = argparse.ArgumentParser(description='Analyze NFL team matchups')
    parser.add_argument('--team1', type=str, default='NYJ', help='First team abbreviation')
    parser.add_argument('--team2', type=str, default='HOU', help='Second team abbreviation')
    args = parser.parse_args()
    
    logging.info(f"Analyzing matchup between {args.team1} and {args.team2}")
    
    # Analyze matchup
    results = analyze_matchup(df, args.team1, args.team2)
    
    # Export to Excel
    export_to_excel(results, args.team1, args.team2)

if __name__ == "__main__":
    main() 